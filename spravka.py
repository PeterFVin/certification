import openpyxl

# theFile = openpyxl.load_workbook('08 Справка ВСЕМАФ.xlsx')
# print(theFile.sheetnames)
# currentSheet = theFile['На регистрацию сертиф.']
# print(currentSheet['D8'].value)
# currentSheet['D8'].value = 123
# theFile.save('table.xlsx')

# theFile = openpyxl.load_workbook('123.xlsx')
# currentSheet = theFile['Лист1']
# a = currentSheet['B2'].value
# b = currentSheet['B3'].value
# print(type(a))

from .certification_many_gui import main_func

a = {
    0: {
        '{{CONTRACT_NUMBER}}': '01-1',
        '{{CONTRACT_YEAR}}': '01-2',
        '{{CONTR_DATE}}': '01-3',
        '{{CONTR_MONTH}}': '01-4',
    },
    1: {
        '{{CONTRACT_NUMBER}}': '02-1',
        '{{CONTRACT_YEAR}}': '02-2',
        '{{CONTR_DATE}}': '02-3',
        '{{CONTR_MONTH}}': '02-4',
    },
}
main_func(a)