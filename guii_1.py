import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook

from func_genitive import genitive, genitive_name, number_to_words
from guii_2 import print_func
from certification_many_gui import main_func

MONTHS = {
    '01': 'январь',
    '02': 'февраль',
    '03': 'март',
    '04': 'апрель',
    '05': 'май',
    '06': 'июнь',
    '07': 'июль',
    '08': 'август',
    '09': 'сентябрь',
    '10': 'октябрь',
    '11': 'ноябрь',
    '12': 'декабрь',
}

var_list = (
    ('CONTRACT_NUMBER', 'B2', 'Введите номер договора, например 19'),
    ('CONTRACT_YEAR', 'B3', 'Введите 2 последние цифры года договора, например 24'),
    ('WORK_NUMBER', 'B4', 'Введите номер работы - если несколько работ, то например 19-01, 19-02 и т.д.24'),
    ('BILL_NUMBER', 'B5', 'Введите номер счета'),
    ('CONTR_DATE', 'B6', 'Введите дату (день) договора, например 01'),
    ('CONTR_MONTH', 'B7', 'CONTR_MONTH', 'Введите месяц договора числом, например 01 если январь'),
    ('BUSINESS_FORM_FULL', 'B8', 'Введите организационно-правовую форму полностью, например общество с ограниченной ответственностью'),
    ('COMPANY_NAME_FULL', 'B9', 'Введите наименование организации (полное)'),
    ('BUSINESS_FORM', 'B10', 'Введите организационно-правовую форму кратко, например ООО'),
    ('COMPANY_NAME', 'B11', 'Введите сокращенное имя компании'),
    ('DIR_LASTNAME', 'B12', 'Введите фамилию директора'),
    ('DIR_FIRSTNAME', 'B13', 'Введите имя директора'),
    ('DIR_SECNAME', 'B14', 'Введите отчество директора'),
    ('GENDER', 'B15', 'Введите пол директора, строго М или Ж'),
    ('CERT_NAME', 'B16', 'Введите наименование сертификата'),
    ('OKPD', 'B17', 'Введите код ОКПД 2'),
    ('STANDART_MAIN', 'B18', 'Введите основной стандарт (указывается в акте отбора)'),
    ('STANDART_SHORT', 'B19', 'Введите перечень стандартов - без пунктов'),
    ('STANDART_FULL', 'B20', 'Введите перечень стандартов с пунктами (если есть отдельные пункты'),
    ('CONTRACT_SUM', 'B21', 'Введите полную сумму договора)'),
    ('CONTRACT_OS_FULL_SUM', 'B22', 'Введите общую сумму оплаты услуг по ОС'),
    ('CONTRACT_OS_SUM', 'B23', 'Введите общую сумму оплаты услуг по ОС по одному сертифицируемому объекту'),
    ('CONTRACT_IL_FULL_SUM', 'B24', 'Введите общую сумму оплаты услуг по ИЛ'),
    ('CONTRACT_IL_SUM', 'B25', 'Введите общую сумму оплаты услуг по ИЛ по одному сертифицируемому объекту'),
    ('CONTRACT_IK_SUM', 'B26', 'Введите сумму договора инспекционного контроля'),
    ('INN', 'B27', 'Введите ИНН организации'),
    ('KPP', 'B28', 'Введите КПП организации'),
    ('JUR_ADDRESS', 'B29', 'Введите юридический адрес организации'),
    ('PHYS_ADDRESS', 'B30', 'Введите физический адрес организации'),
    ('RAS_ACCOUNT', 'B31', 'Введите расчетный счет организации'),
    ('BANK_NAME', 'B32', 'Введите имя банка организации'),
    ('KORR_ACCOUNT', 'B33', 'Введите корреспондентский счет организации'),
    ('BIK', 'B34', 'Введите БИК банка организации'),
    ('OGRN', 'B35', 'Введите ЫЫЫЫЫЫЫЫЫЫЫЫЫЫ организации'),
    ('TEL', 'B36', 'Введите телефон организации'),
    ('E-MAIL', 'B37', 'Введите e-mail организации'),
    ('EXPERT_LASTNAME', 'B38', 'Введите фамилию эксперта по сертификации'),
    ('EXPERT_FIRSTNAME_SHORT', 'B39', 'Введите сокращенное имя эксперта по сертификации (с точкой) например П.'),
    ('EXPERT_SECNAME_SHORT', 'B40', 'Введите сокращенное отчество эксперта по сертификации (с точкой) например Ф.'),
    ('EXPERT_REG_NUMBER', 'B41', 'Введите номер эксперта в реестре МСС'),
    ('IL_NAME', 'B42', 'Введите наименование ИЛ'),
    ('IL_REG_NUMBER', 'B43', 'Введите номер ИЛ в реестре МСС'),
    ('IL_EXPIRE_DATE', 'B44', 'Введите дату окончания действия аккредитация ИЛ'),
    ('ISSUE_DECISION_DATE', 'B45', 'Введите дату (день) решения о выдаче сертификата, например 01'),
    ('ISSUE_DECISION_MONTH', 'B46', 'Введите месяц решения о выдаче сертификата - числом, например 01 если январь'),
    ('ISSUE_DECISION_YEAR', 'B47', 'Введите год решения о выдаче сертификата - четыре цифры'),
    ('CERTIFICARTE_START_DATE', 'B48', 'Введите дату (день) начала действия сертификата, например 01'),
    ('CERTIFICARTE_START_MONTH', 'B49', 'Введите Введите месяц начала действия сертификата - числом, например 01 если январь'),
    ('CERTIFICARTE_START_YEAR', 'B50', 'Введите год начала действия сертификата - четыре цифры'),
    ('CERTIFICARTE_DURATION', 'B51', 'Введите срок действия сертификата, например 3'),
    ('SAMPLE_ACT_DATE', 'B52', 'Введите дату (день) акта отбора образцов, например 01'),
    ('SAMPLE_ACT_MONTH', 'B53', 'Введите месяц акта отбора образцов - числом, например 01 если январь'),
    ('SAMPLE_ACT_YEAR', 'B54', 'Введите год акта отбора образцов - четыре цифры'),
    ('PRODUCTION_CREATED_QUARTER', 'B55', 'Введите квартал изготовления продукции, например IV'),
    ('PRODUCTION_CREATED_YEAR', 'B56', 'Введите год изготовления продукции - четыре цифры'),
    ('PRODUCTION_SAMPLES', 'B57', 'Введите описание образцов продукции для испытаний'),
    ('PROTOCOL_DATE', 'B58', 'Введите дату (день) протокола, например 01'),
    ('PROTOCOL_MONTH', 'B59', 'Введите месяц протокола - числом, например 01 если январь'),
    ('PROTOCOL_YEAR', 'B60', 'Введите год протокола - четыре цифры'),
    ('PROTOCOL_NUMBER', 'B61', 'Введите номер протокола'),
    ('TEST_GOSTS', 'B62', 'Введите стандарты методики испытаний'),
    ('TEST_START_DATE', 'B63', 'Введите дату (день) начала сертификационных испытаний, например 01'),
    ('TEST_START_MONTH', 'B64', 'Введите месяц начала сертификационных испытаний - числом, например 01 если январь'),
    ('TEST_START_YEAR', 'B65', 'Введите год начала сертификационных испытаний - четыре цифры'),
    ('SAMPLES_MARK', 'B66', 'Введите маркировку образцов для испытаний'),
    ('TESTER_NAME', 'B67', 'Введите Фамилию И.О. проводившего испытания'),
    ('PROD_ANALYSE_DATE', 'B68', 'Введите дату (день) анализа производства, например 01'),
    ('PROD_ANALYSE_MONTH', 'B69', 'Введите месяц анализа производства - числом, например 01 если январь'),
    ('PROD_ANALYSE_YEAR', 'B70', 'Введите год анализа производства - четыре цифры'),
)

# Function to import data from Excel
def import_data():
    # Open a file dialog to select the Excel file
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls")])
    if not file_path:
        return  # If no file is selected, do nothing


    # Load the workbook and get sheet names
    workbook = load_workbook(file_path)
    sheet_names = workbook.sheetnames
    # Clear existing sheets
    global sheets
    sheets = {}
    for tab in notebook.tabs():
        notebook.forget(tab)
    # Loop through each sheet in the Excel file
    for idx, sheet_name in enumerate(sheet_names, start=1):
        worksheet = workbook[sheet_name]
        # Create a new sheet in the Tkinter notebook
        create_sheet()
        
        
        
        # Populate data from the Excel sheet
        entries = sheets[f"Серт {idx}"]['entries']
        for i in range(len(var_list)):
            entries[var_list[i][0]].delete(0, tk.END)
            entries[var_list[i][0]].insert(0, worksheet[var_list[i][1]].value)
            
            # entries['contract_year'].delete(0, tk.END)
            # entries['contract_year'].insert(0, worksheet["A2"].value)
            
            # entries['company_name'].delete(0, tk.END)
            # entries['company_name'].insert(0, worksheet["A3"].value)


def validate_fields(entries):
    """
    Validate the fields of a single sheet. Returns a list of error messages.
    """
    errors = []

    # Validate CONTRACT_NUMBER
    contract_number = entries[var_list[0][0]].get()
    if not contract_number.isdigit() or int(contract_number) > 100:
        errors.append("Field 'CONTRACT_NUMBER' should be an integer <= 100.")

    # Validate CONTRACT_YEAR
    contract_year = entries['CONTRACT_YEAR'].get()
    if not contract_year.isdigit():
        errors.append("Field 'CONTRACT_YEAR' should be an integer.")

    # # Validate COMPANY_NAME
    # company_name = entries['company_name'].get()
    # if not all(char in "AB" for char in company_name):
    #     errors.append("Field 'COMPANY_NAME' should contain only letters 'A' or 'B'.")

    return errors


root = tk.Tk()
root.title("Document Generator")
root.geometry("500x600")  # Set a fixed window size

# Create a container frame for the canvas and scrollbar
container = tk.Frame(root)
container.pack(fill="both", expand=True)

# Create a canvas for the notebook
canvas = tk.Canvas(container)
canvas.pack(side="left", fill="both", expand=True)

# Add a vertical scrollbar to the canvas
scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
scrollbar.pack(side="right", fill="y")
canvas.configure(yscrollcommand=scrollbar.set)

# Create a frame inside the canvas to hold the notebook
frame = tk.Frame(canvas)
canvas.create_window((0, 0), window=frame, anchor="nw")

# Function to adjust the scroll region
def on_frame_configure(event):
    canvas.configure(scrollregion=canvas.bbox("all"))

frame.bind("<Configure>", on_frame_configure)

# Track sheet frames and available numbers
sheets = {}
available_numbers = []

def on_ok():
    """
    Validation, Collect data from all sheets and process it.
    """
    
    all_errors = []

    # Iterate through all sheets and validate their fields
    for sheet_name, sheet_data in sheets.items():
        entries = sheet_data['entries']
        errors = validate_fields(entries)
        if errors:
            all_errors.append(f"Errors in {sheet_name}:")
            all_errors.extend(errors)

    # If there are errors, show them in a messagebox
    if all_errors:
        messagebox.showerror("Validation Errors", "\n".join(all_errors))
        return

    variables = {}
    variables['CERT_NAME'] = []
    variables['STANDART_FULL'] = []
    variables['OKPD'] = []
    counter = 0
    documents_single = {
        0: flag1_var.get(),
        1: flag2_var.get(),
        2: flag3_var.get(),
        3: flag4_var.get(),
        4: flag5_var.get(),
        5: flag6_var.get()
    }
    documents_many = {
        0: flag7_var.get(),
        1: flag8_var.get(),
        2: flag9_var.get(),
        3: flag10_var.get(),
        4: flag11_var.get(),
        5: flag12_var.get(),
        6: flag13_var.get()
    }
    spravka = flag14_var.get()

    for sheet_name, sheet_frame in sheets.items():
        counter += 1
        variables[counter] = {}
        entries = sheet_frame['entries']
        variables[counter]['worksheets_count'] = len(sheets)
        # Эти три списка в словаре - для добавления нескольких строк продукции в договорах и заявке
        # Работают в строке - if template_file_path == '00 Подписанный Дог серт Исходник.docx'...
        for i in range(len(var_list)):
            variables[counter]['{{' + var_list[i][0] + '}}'] = entries[var_list[i][0]].get()
            if var_list[i][0] in ['CERT_NAME', 'OKPD', 'STANDART_FULL']:
                variables[var_list[i][0]].append(variables[counter]['{{' + var_list[i][0] + '}}'])

        #  Для распоряжения. Ограничение максимального дня - 28, чтобы не получилось 29.02
        variables[counter]['{{CONTR_DATE<29}}'] = variables[counter]['{{CONTR_DATE}}'] if int(variables[counter]['{{CONTR_DATE}}']) < 29 else '28'
        #  Если месяц договора декабрь, то месяц окончания работ - январь, чтобы не получилось месяц окончания - 13
        variables[counter]['{{CONTR_MONTH+1}}'] = '01' if variables[counter]['{{CONTR_MONTH}}'] == '12' else str(int(variables[counter]['{{CONTR_MONTH}}']) + 1)  # Для указания месяца проведения работ.
        variables[counter]['{{CONTR_MONTH_WORD_GEN}}'] = genitive(MONTHS[variables[counter]['{{CONTR_MONTH}}']])
        variables[counter]['{{CONTR_YEAR}}'] = '20' + variables[counter]['{{CONTRACT_YEAR}}']
        #  Равен году договора, а если месяц договора декабрь, то год договора + 1
        variables[counter]['{{CONTR_YEAR+1}}'] = variables[counter]['{{CONTR_YEAR}}'] if variables[counter]['{{CONTR_MONTH}}'] != '12' else str(int(variables[counter]['{{CONTR_YEAR}}']) + 1)
        variables[counter]['{{DIR_FIRSTNAME_SHORT}}'] = variables[counter]['{{DIR_FIRSTNAME}}'][:1] + '.'
        variables[counter]['{{DIR_SECNAME_SHORT}}'] = variables[counter]['{{DIR_SECNAME}}'][:1] + '.'
        variables[counter]['{{DIR_LASTNAME_GEN}}'] = genitive_name('LASTNAME', variables[counter]['{{GENDER}}'], variables[counter]['{{DIR_LASTNAME}}'])  # Фамилия
        variables[counter]['{{DIR_FIRSTNAME_GEN}}'] = genitive_name('FIRSTNAME', variables[counter]['{{GENDER}}'], variables[counter]['{{DIR_FIRSTNAME}}'])  # Имя
        variables[counter]['{{DIR_SECNAME_GEN}}'] = genitive_name('MIDDLENAME', variables[counter]['{{GENDER}}'], variables[counter]['{{DIR_SECNAME}}'])  # Отчество
        variables[counter]['{{CONTRACT_SUM_WORDS}}'] = number_to_words(variables[counter]['{{CONTRACT_SUM}}'])
        variables[counter]['{{ISSUE_DECISION_MONTH_GEN}}'] = genitive(MONTHS[variables[counter]['{{ISSUE_DECISION_MONTH}}']])
        variables[counter]['{{SAMPLE_ACT_MONTH_GEN}}'] = genitive(MONTHS[variables[counter]['{{SAMPLE_ACT_MONTH}}']])
        variables[counter]['{{PROTOCOL_MONTH_WORD_GEN}}'] = genitive(MONTHS[variables[counter]['{{PROTOCOL_MONTH}}']])
        variables[counter]['{{CERTIFICARTE_START_MONTH_WORD}}'] = genitive(MONTHS[variables[counter]['{{CERTIFICARTE_START_MONTH}}']])
        variables[counter]['{{CERTIFICARTE_START_YEAR+1}}'] = str(int(variables[counter]['{{CONTR_YEAR}}']) + 1)
        variables[counter]['{{CERTIFICARTE_START_YEAR+2}}'] = str(int(variables[counter]['{{CONTR_YEAR}}']) + 2)
        variables[counter]['{{CERTIFICARTE_START_YEAR+N}}'] = str(int(variables[counter]['{{CONTR_YEAR}}']) + int(variables[counter]['{{CERTIFICARTE_DURATION}}']))

    print(variables[counter]['{{CONTR_YEAR}}'])
    value = variables[1]['{{CONTRACT_IK_SUM}}']

    # # Для Python 3.11
    # match int(variables[1]['{{CERTIFICARTE_DURATION}}']):
    #     case 4:
    #         variables[1]['{{IK_ADD_SUM}}'] = f"\n\tIII этап\t\t{value} руб. 00 коп."
    #     case 5:
    #         variables[1]['{{IK_ADD_SUM}}'] = f"\n\tIII этап\t\t{value} руб. 00 коп.\n\tVI этап\t\t{value} руб. 00 коп."
    #     case _:
    #         variables[1]['{{IK_ADD_SUM}}'] = ''

    var1 = int(variables[1]['{{CERTIFICARTE_DURATION}}'])
    if var1 == 4:
        variables[1]['{{IK_ADD_SUM}}'] = f"\n\tIII этап\t\t{value} руб. 00 коп."
    elif var1 == 5:
            variables[1]['{{IK_ADD_SUM}}'] = f"\n\tIII этап\t\t{value} руб. 00 коп.\n\tVI этап\t\t{value} руб. 00 коп."
    else:
        variables[1]['{{IK_ADD_SUM}}'] = ''


            # variables[counter]['{{CONTRACT_NUMBER}}'] = entries[var_list[0][0]].get()
            # variables[counter]['{{CONTRACT_YEAR}}'] = entries['contract_year'].get()
            # variables[counter]['{{COMPANY_NAME}}'] = entries['company_name'].get()

    # print_func(variables, documents)
    main_func(variables, documents_single, documents_many, spravka)
    messagebox.showinfo('Программа выполнена!', 'Программа отработала успешно, все документы созданы!')

def create_sheet():
    # Determine the next available sheet number
    sheet_number = len(sheets) + 1
    sheet_name = f"Серт {sheet_number}"

    # Create a new frame for each sheet
    frame = ttk.Frame(notebook)
    notebook.add(frame, text=sheet_name)

    # Dictionary to store Entry widgets for easy access
    entries = {}

    # Retrieve data from the last sheet if it exists
    last_sheet_name = f"Серт {sheet_number - 1}"
    last_entries = sheets.get(last_sheet_name, {}).get('entries', {})

    # Add labels, entry fields, and tip text to the sheet
    for i in range(len(var_list)):
        tk.Label(frame, text=f'{var_list[i][0]}:').grid(row=i, column=0, padx=10, pady=5, sticky="e")
        entries[var_list[i][0]] = tk.Entry(frame)
        entries[var_list[i][0]].grid(row=i, column=1, padx=10, pady=5)
        if last_entries:
            entries[var_list[i][0]].insert(0, last_entries[var_list[i][0]].get())
        tk.Label(frame, text=var_list[i][2]).grid(row=i, column=2, padx=10, pady=5, sticky="w")
    


    # tk.Label(frame, text="CONTRACT_YEAR:").grid(row=1, column=0, padx=10, pady=5, sticky="e")
    # entries['contract_year'] = tk.Entry(frame)
    # entries['contract_year'].grid(row=1, column=1, padx=10, pady=5)
    # if last_entries:
    #     entries['contract_year'].insert(0, last_entries['contract_year'].get())
    # tk.Label(frame, text="Enter a valid year (e.g., 2024)").grid(row=1, column=2, padx=10, pady=5, sticky="w")

    # tk.Label(frame, text="COMPANY_NAME:").grid(row=2, column=0, padx=10, pady=5, sticky="e")
    # entries['company_name'] = tk.Entry(frame)
    # entries['company_name'].grid(row=2, column=1, padx=10, pady=5)
    # if last_entries:
    #     entries['company_name'].insert(0, last_entries['company_name'].get())
    # tk.Label(frame, text="Enter 'A' or 'B' only").grid(row=2, column=2, padx=10, pady=5, sticky="w")

    # Store the frame and its entries in the sheets dictionary
    sheets[sheet_name] = {'frame': frame, 'entries': entries}


def remove_sheet():
    if notebook.tabs():  # Check if there are any tabs
        # Get the last tab
        last_tab = notebook.tabs()[-1]
        # Find the name associated with the last tab ID
        for name, sheet_data in list(sheets.items()):
            if str(sheet_data['frame']) == last_tab:
                # Extract sheet number from the name and add it back to available_numbers
                sheet_number = int(name.split()[-1])
                available_numbers.append(sheet_number)

                # Remove the tab from the notebook and delete it from the dictionary
                notebook.forget(sheet_data['frame'])
                del sheets[name]
                break

def set_all_flags():
    flag1_var.set(True)
    flag2_var.set(True)
    flag3_var.set(True)
    flag4_var.set(True)
    flag5_var.set(True)
    flag6_var.set(True)
    flag7_var.set(True)
    flag8_var.set(True)
    flag9_var.set(True)
    flag10_var.set(True)
    flag11_var.set(True)
    flag12_var.set(True)
    flag13_var.set(True)
    flag14_var.set(True)

def unset_all_flags():
    flag1_var.set(False)
    flag2_var.set(False)
    flag3_var.set(False)
    flag4_var.set(False)
    flag5_var.set(False)
    flag6_var.set(False)
    flag7_var.set(False)
    flag8_var.set(False)
    flag9_var.set(False)
    flag10_var.set(False)
    flag11_var.set(False)
    flag12_var.set(False)
    flag13_var.set(False)
    flag14_var.set(False)

# Create the Notebook widget inside the frame
notebook = ttk.Notebook(frame)
notebook.pack(pady=10, padx=10, fill="both", expand=True)

# Initial setup: Add one default sheet
create_sheet()

# Buttons to add, remove sheets, and a global OK button
button_frame = tk.Frame(root)
button_frame.pack(pady=10)

# Initialize flag variables
flag1_var = tk.BooleanVar(value=False)  # Default to False
flag2_var = tk.BooleanVar(value=False)  # Default to False
flag3_var = tk.BooleanVar(value=False)  # Default to False
flag4_var = tk.BooleanVar(value=False)  # Default to False
flag5_var = tk.BooleanVar(value=False)  # Default to False
flag6_var = tk.BooleanVar(value=False)  # Default to False
flag7_var = tk.BooleanVar(value=False)  # Default to False
flag8_var = tk.BooleanVar(value=False)  # Default to False
flag9_var = tk.BooleanVar(value=False)  # Default to False
flag10_var = tk.BooleanVar(value=False)  # Default to False
flag11_var = tk.BooleanVar(value=False)  # Default to False
flag12_var = tk.BooleanVar(value=False)  # Default to False
flag13_var = tk.BooleanVar(value=False)  # Default to False
flag14_var = tk.BooleanVar(value=False)  # Default to False

# Add a frame for flags and buttons
control_frame = tk.Frame(root)
control_frame.pack(pady=10, fill="x")  # Ensure it spans the width of the window

# Add flags in the first row of the control frame
flag1_checkbox = tk.Checkbutton(control_frame, text="Подписанный договор", variable=flag1_var)
flag1_checkbox.grid(row=0, column=0, padx=5, pady=5, sticky="w")

flag2_checkbox = tk.Checkbutton(control_frame, text="Договор сертификация", variable=flag2_var)
flag2_checkbox.grid(row=0, column=1, padx=5, pady=5, sticky="w")

flag3_checkbox = tk.Checkbutton(control_frame, text="Счет", variable=flag3_var)
flag3_checkbox.grid(row=0, column=2, padx=5, pady=5, sticky="w")

flag4_checkbox = tk.Checkbutton(control_frame, text="Договор инспекция", variable=flag4_var)
flag4_checkbox.grid(row=0, column=3, padx=5, pady=5, sticky="w")

flag5_checkbox = tk.Checkbutton(control_frame, text="Акт закрытия", variable=flag5_var)
flag5_checkbox.grid(row=0, column=4, padx=5, pady=5, sticky="w")

flag6_checkbox = tk.Checkbutton(control_frame, text="Заявка", variable=flag6_var)
flag6_checkbox.grid(row=0, column=5, padx=5, pady=5, sticky="w")

flag7_checkbox = tk.Checkbutton(control_frame, text="Распоряжение", variable=flag7_var)
flag7_checkbox.grid(row=1, column=0, padx=5, pady=5, sticky="w")

flag8_checkbox = tk.Checkbutton(control_frame, text="Акт отбора", variable=flag8_var)
flag8_checkbox.grid(row=1, column=1, padx=5, pady=5, sticky="w")

flag9_checkbox = tk.Checkbutton(control_frame, text="Протокол", variable=flag9_var)
flag9_checkbox.grid(row=1, column=2, padx=5, pady=5, sticky="w")

flag10_checkbox = tk.Checkbutton(control_frame, text="Анализ производства", variable=flag10_var)
flag10_checkbox.grid(row=1, column=3, padx=5, pady=5, sticky="w")

flag11_checkbox = tk.Checkbutton(control_frame, text="Заключение", variable=flag11_var)
flag11_checkbox.grid(row=1, column=4, padx=5, pady=5, sticky="w")

flag12_checkbox = tk.Checkbutton(control_frame, text="Решение о выдаче", variable=flag12_var)
flag12_checkbox.grid(row=1, column=5, padx=5, pady=5, sticky="w")

flag13_checkbox = tk.Checkbutton(control_frame, text="Макет сертификата", variable=flag13_var)
flag13_checkbox.grid(row=1, column=6, padx=5, pady=5, sticky="w")

flag14_checkbox = tk.Checkbutton(control_frame, text="Справка", variable=flag14_var)
flag14_checkbox.grid(row=1, column=7, padx=5, pady=5, sticky="w")


# Add buttons in the second row of the control frame
add_button = tk.Button(control_frame, text="Add Frame", command=create_sheet)
add_button.grid(row=3, column=0, padx=5, pady=5)

remove_button = tk.Button(control_frame, text="Remove Frame", command=remove_sheet)
remove_button.grid(row=3, column=1, padx=5, pady=5)

import_button = tk.Button(control_frame, text="Import Data", command=import_data)
import_button.grid(row=3, column=2, padx=5, pady=5)

ok_button = tk.Button(control_frame, text="OK", command=on_ok)
ok_button.grid(row=3, column=3, padx=5, pady=5)

# Add Set/Unset buttons in the second row of the control frame
set_flags_button = tk.Button(control_frame, text="Set All Flags", command=set_all_flags)
set_flags_button.grid(row=2, column=0, padx=5, pady=5)

unset_flags_button = tk.Button(control_frame, text="Unset All Flags", command=unset_all_flags)
unset_flags_button.grid(row=2, column=1, padx=5, pady=5)

# # Configure column weights for proper alignment
# control_frame.columnconfigure(0, weight=1)
# control_frame.columnconfigure(1, weight=1)
# control_frame.columnconfigure(2, weight=1)
# control_frame.columnconfigure(3, weight=1)

# Run the main event loop
root.mainloop()
